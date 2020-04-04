import os
from os.path import exists
from time import sleep


os.environ.setdefault("DJANGO_SETTINGS_MODULE", "CRM.settings")
os.environ.setdefault('C_FORCE_ROOT', 'true')
from CRM.Core.Signals import service_update_request
from CRM.Core.CRMConfig import read_config
from celery.schedules import crontab
from django.db import transaction
from CRM.models import GiftHistory, UserServiceGroup, NotifySettings, ServiceGroups, UserIPStatic, \
    IBSUserInfo, IPPool, ResellerProfile, EquipmentOrder, EquipmentUnknownCondition, InvoiceChargeState
from CRM.Tools.DateParser import parse_date_from_str_to_julian, parse_date_from_str
from CRM.templatetags.DateConverter import convert_date
from datetime import timedelta, datetime
from CRM.Tools.SendNotification import __render__, send_email, send_inbox, send_text_message
from CRM.Tools.Validators import validate_integer, validate_ip

from CRM.Core.CRMUserUtils import assign_all_users_to_price_list, get_reseller_user_ids, get_reseller_profile_data
from celery import Celery
from django.contrib.auth.models import Group, User
from django.core.cache import cache
from CRM.IBS.Manager import IBSManager
from CRM.IBS.Users import IBSUserManager
from CRM.Processors.PTools.Utility import create_user_from_ibs, convert_credit, update_user_services
from CRM.Processors.PTools.Core.Charge.Service.IPStatic import de_configure_user_static_ip
from openpyxl import load_workbook
from khayyam.jalali_date import JalaliDate

# from CRM.models import BaseConfig, UsersPriceList, PriceList, NotifySettings, Factor, GiftHistory

__author__ = 'Amir'

app = Celery('Jobs', broker='redis://localhost')

app.conf.update(CELERYBEAT_SCHEDULE={
    'clear_invoice_crontab': {
        'task': 'clear_invoice',
        'schedule': crontab(minute=0, hour=9)
    },
    'send_gift_crontab': {
        'task': 'send_gift_by_date',
        'schedule': crontab(minute=15, hour=9)
    },
    'check_ip_release_crontab': {
        'task': 'check_for_release_ip',
        'schedule': crontab(minute=30, hour=9)
    },
    'send_expire_notification_crontab': {
        'task': 'send_ip_expire_notification',
        'schedule': crontab(minute=0, hour=10)
    },
    'expire_user_ip_crontab': {
        'task': 'expire_user_ip',
        'schedule': crontab(minute=0, hour=11)
    },
    'reseller_profit': {
        'task': 'calculate_reseller_profit',
        'schedule': crontab(minute=0, hour=7)
    },
    'unknown_condition_check': {
        'task': 'return_unknown_condition_equipment',
        'schedule': crontab(minute=0, hour=7)
    }
}, CELERY_TIMEZONE='Asia/Tehran', CELERYBEAT_MAX_LOOP_INTERVAL=180)


@app.task(name='return_unknown_condition_equipment')
def return_unknown_condition_equipment():
    dt = datetime.today() - timedelta(days=2)
    data = EquipmentOrder.objects.filter(receiver__isnull=False, fk_equipment_in_use_order=None,
                                         receive_date__lt=dt)
    for d in data:
        x = EquipmentUnknownCondition()
        x.order_id = d.pk
        x.save()


@app.task(name='calculate_reseller_profit')
def calculate_reseller_profit():
    now = JalaliDate.today()
    if now.day > 1:
        return
    resellers = ResellerProfile.objects.all()
    for f in resellers:
        users = get_reseller_user_ids(f.user_id)
        data = get_reseller_profile_data(users, f.user_id)
        p = data.get('total_service_price')
        if p is None:
            p = 0
        s = data.get('total_package_price')
        if s is None:
            s = 0
        t = s + p
        f.old_price = f.profit_price
        f.profit_price += t
        f.save()


@app.task
def import_ip_statics_job():
    cache.set("IBS_IMPORTING", 1)
    try:
        ibm = IBSManager()
        ibu = IBSUserManager(ibm)
        ibs_users = IBSUserInfo.objects.all()
        for u in ibs_users:
            ip = ibu.get_user_ip_static(u.ibs_uid)
            if validate_ip(ip):
                if IPPool.objects.filter(ip=ip).exists():
                    pool = IPPool.objects.get(ip=ip)
                else:
                    pool = IPPool()
                    pool.ip = ip
                    pool.save()
                    print '[IP-STATIC] ADDRESS ADDED : %s' % ip
                if not UserIPStatic.objects.filter(user=u.user_id).exists():
                    user_ip = UserIPStatic()
                    user_ip.release_date = datetime.today() + timedelta(days=30)
                    user_ip.expire_date = datetime.today() + timedelta(days=15)
                    user_ip.start_date = datetime.today()
                    user_ip.service_period = 1
                else:
                    user_ip = UserIPStatic.objects.get(user=u.user_id)
                user_ip.ip = pool
                user_ip.user = User.objects.get(pk=u.user_id)
                user_ip.is_reserved = False
                user_ip.request_time = datetime.today()
                user_ip.save()
                # if StaticIPRequest.objects.filter(user=u.user_id).exists():
                #     rq = StaticIPRequest.objects.get(user=u.user_id)
                # else:
                #     rq = StaticIPRequest()
                #     rq.end_date = datetime.today() + timedelta(days=15)
                #     rq.request_time = datetime.today()
                #     rq.service_period = 1
                #     rq.start_date = datetime.today()
                # rq.user = User.objects.get(pk=u.user_id)
                # rq.address = ip
                # rq.save()
                print '[IP-STATIC] IP ASSIGNED TO USER %s %s' % (ip, u.user.username)
        if exists('/var/CRM/Docs/None/IP.xlsx'):
            wb = load_workbook('/var/CRM/Docs/None/IP.xlsx')
            act = wb.get_active_sheet()
            for i in range(2, act.max_row):
                row_id1 = 'A%s' % i
                row_id2 = 'B%s' % i
                uid = act[row_id1].value
                if not validate_integer(uid):
                    break
                exp = act[row_id2].value
                date0 = exp.split('.')
                date1 = '13%s-%s-%s' % (date0[2], date0[1], date0[0])
                # print(date1)
                cor_date = parse_date_from_str(date1)
                # print cor_date
                # cor_date.hour = 12
                # cor_date.minute = 0
                if UserIPStatic.objects.filter(user__fk_ibs_user_info_user__ibs_uid=int(uid)).exists():
                    user_ip_data = UserIPStatic.objects.get(user__fk_ibs_user_info_user__ibs_uid=int(uid))
                    user_ip_data.expire_date = cor_date
                    user_ip_data.save()
                    # StaticIPRequest.end_date
                    # req = StaticIPRequest.objects.filter(user__fk_ibs_user_info_user__ibs_uid=int(uid)).last()
                    # req.end_date = cor_date
                    # req.save()
    except Exception as e:
        print e.message
    finally:
        cache.set("IBS_IMPORTING", 0)


@app.task
def kill_user_by_request(user_id, request):
    from CRM.Core.ServiceManager import Utils
    Utils.kill_user_by_request(user_id, request)


@app.task(name='expire_user_ip')
def expire_user_ip():
    ips = UserIPStatic.objects.filter(expire_date__lt=datetime.today().date(), is_free=False, is_reserved=False)
    print('%s ips to expire!' % ips.count())
    for i in ips:
        de_configure_user_static_ip(i.user_id)
        send_from_template(i.user_id, 18, sti=i.ip.ip)
        kill_user_by_request.delay(i.user_id, None)


@app.task(name='send_ip_expire_notification')
def send_ip_expire_notification():
    ips = UserIPStatic.objects.filter(expire_date__lt=datetime.today() + timedelta(days=3), is_free=False,
                                      is_reserved=False)
    print '%s ips will expire in next 3 days' % ips.count()
    for i in ips:
        send_from_template(i.user_id, 16, exp=convert_date(i.expire_date))


@app.task(name='check_for_release_ip')
def check_for_release_ip():
    ips = UserIPStatic.objects.filter(release_date__lt=datetime.today(), is_reserved=True, is_free=False)
    print 'found %s ips to release' % ips.count()
    for i in ips:
        i.delete()


@app.task(name='send_gift_by_date')
def send_gift_for_users_by_date():
    marriage_users = User.objects.filter(fk_user_profile_user__marriage_date_day=datetime.today().date().day,
                                         fk_user_profile_user__marriage_date_month=datetime.today().date().month)
    birth_users = User.objects.filter(fk_user_profile_user__birth_date_month=datetime.today().date().month,
                                      fk_user_profile_user__birth_date_day=datetime.today().day)
    birth_package = read_config(name='gift_birthday_package', default=0)
    birth_service = read_config(name='gift_birthday_service', default=0)
    birth_add_limited = False
    marriage_package = read_config(name='gift_marriage_package', default=0)
    marriage_service = read_config(name='gift_marriage_service', default=0)
    marriage_add_limited = False
    # Calling Them
    send_gift_for_users(birth_users, int(birth_service), int(birth_package), birth_add_limited)
    send_gift_for_users(marriage_users, int(marriage_service), int(marriage_package), marriage_add_limited)


@app.task
def send_gift_for_users(users, add_days, add_package, add_days_for_limited=False):
    # assert isinstance(users, User.objects)
    for u in users:
        # assert isinstance(u, User)
        if not u.is_active:
            continue
        if not u.fk_user_current_service_user.exists():
            continue
        user_service = u.fk_user_current_service_user.get()
        if not u.fk_ibs_user_info_user.exists():
            continue
        ibi = u.fk_ibs_user_info_user.get().ibs_uid
        if not user_service.is_active:
            continue
        ibs = IBSManager()
        gh = GiftHistory()
        gh.target_user = u
        gh.extended_days = int(add_days)
        gh.extended_package = int(add_package)
        gh.save()
        if add_days > 0:
            if user_service.service_property.initial_package > 0 and add_days_for_limited:
                ibs.set_expire_date_by_uid(ibi, add_days)
            if user_service.service_property.initial_package == 0:
                ibs.set_expire_date_by_uid(ibi, add_days)
        if add_package > 0 and user_service.service_property.initial_package > 0:
            ibs.change_credit_by_uid(add_package, ibi)
            add_package = convert_credit(add_package)
            trm = convert_credit(ibs.get_user_credit_by_user_id(ibi))
        else:
            trm = None
        kw = {'name': '', 'ta': add_package,
              'exp': convert_date(parse_date_from_str_to_julian(str(ibs.get_expire_date_by_uid(ibi)))),
              'ads': add_days,
              'trm': trm}
        send_from_template(u.pk, 17, **kw)


@app.task(name='clear_invoice')
def clear_unused_invoices():
    from CRM.models import Invoice
    current_date_time = datetime.today().date() - timedelta(days=2)
    transaction.set_autocommit(False)
    try:
        res = Invoice.objects.filter(create_time__lt=current_date_time, is_paid=False)
        for i in res:
            print i.pk
            i.is_deleted = True
            i.save()
        transaction.commit()
    except Exception as e:
        print e.message
        transaction.rollback()
    finally:
        transaction.set_autocommit(True)


@app.task
def assign_current_users_to_service_group(pid):
    cache.set('SETTING_USERS', 1)
    assign_all_users_to_price_list(pid)
    cache.set('SETTING_USERS', 0)


@app.task
def import_ibs_users_job(default_price):
    cache.set("IBS_IMPORTING", 1)
    try:
        ibs = IBSManager()
        users = ibs.get_all_users()
        # l.debug("Found %s users" % len(users))
        default_group = read_config(name='default_customer_group', default=1)
        group = Group.objects.get(pk=default_group)
        for uid in users:
            cid = create_user_from_ibs(uid)
            if not cid[0]:
                continue
            # l.info("Adding user to internet group")
            u = User.objects.get(pk=cid[2])
            if not UserServiceGroup.objects.filter(user=u.pk).exists():
                up = UserServiceGroup()
                up.user = u
                up.service_group = ServiceGroups.objects.get(pk=default_price)
                up.save()
            if u.groups.filter(pk=default_group).exists():
                # l.info("User is in group!")
                continue
            try:
                u.groups.add(group)
                # l.info("User added to internet group : %s" % uid)
            except Exception as e:
                # l.error("Unable to add %s to internet group" % uid)
                if e.args:
                    print(" ".join([str(a) for a in e.args]))
                else:
                    print(e.message)
                    # return redirect(reverse(search_users))
    except Exception as e:
        if e.args:
            print (" ".join([str(a) for a in e.args]))
        else:
            print(e.message)
            # return render(request, 'errors/ServerError.html')
    cache.set("IBS_IMPORTING", 0)


@app.task
def send_from_template(user_id, code_id, **kwargs):
    if not validate_integer(user_id):
        return False
    if not validate_integer(code_id):
        return False
    try:
        n = NotifySettings.objects.get(code_id=code_id)
        u = User.objects.get(pk=user_id)
        params = {'username': u.username,
                  'id': u.pk,
                  'name': u.first_name,  # User First name
                  'phone': u.fk_user_profile_user.telephone,  # User telephone
                  'mobile': u.fk_user_profile_user.mobile,  # User Mobile Number
                  'cdt': datetime.today(),  # Current Date Time
                  }
        params.update(kwargs)
        if n.email_enabled:
            res = __render__(n.mail_text, params)
            send_email(res, u.email, 'CRM', user_id, (code_id == 4 or code_id == 6))
        if n.inbox_enabled:
            res = __render__(n.inbox_text, params=params)
            send_inbox(user_id, res, (code_id == 4 or code_id == 6))
        if n.sms_enabled:
            res = __render__(n.sms_text, params)
            send_text_message(user_id, res, (code_id == 4 or code_id == 6))
        return True
    except Exception as e:
        print e.message
        return False


@app.task
def check_for_recharge(invoice):
    if not invoice:
        return
    if invoice.fk_invoice_charge_state_invoice.exists():
        return
    check_time = 0
    while check_time < 5:
        sleep(5)
        if InvoiceChargeState.objects.filter(invoice=invoice.pk).exists():
            return
        check_time += 1
    service_update_request.send_robust('InvoiceChargeChecker', invoice=invoice)
